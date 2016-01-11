from openpyxl import load_workbook
from openpyxl import worksheet
##openpyxl.worksheet.views.Selection

from collections import Counter
from datetime import date as dt

import sys
sys.path.append('..\\')
from MyShiftTime import MyShiftTime


Filename = '展顥訂便當R.xlsx'
cells = [3,4,5,6,7] + [12,13,14,15,16]


def CreateTodayXlsx():
    MST = MyShiftTime()

    LogFilename = MST.IsOrderAvbl()['FileName']


    order = dict()
    with open('..\\' + LogFilename + '.txt') as f:
        for line in f:
##            print(line)
            username = line.strip().split(', ')[3]
            cellname = line.strip().split(', ')[4]
            order[username] = cellname

    cnt = Counter(order.values())


    ## write out xlsx
    wb = load_workbook(Filename)

    sheetnames = wb.get_sheet_names()
    sheetname = sheetnames[MST.weekday]
    sheet = wb[sheetname]

    wb.active = wb.get_index(sheet)
    
    sheet['A1'] = '{0.month}/{0.day} 展顥'.format(dt.today())
    for i in cnt:
        cell = i.replace('A','B')

        if cell !='0':
            sheet[cell] = cnt[i]

        
    NewFileName = "展顥訂便當_{0.month:02d}{0.day:02d}.xlsx".format(dt.today())
    wb.save(NewFileName)

    return NewFileName


CreateTodayXlsx()
