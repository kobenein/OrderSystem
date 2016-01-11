from openpyxl import load_workbook
from MyShiftTime import MyShiftTime


Filename = 'XLSX\展顥訂便當R.xlsx'
cells = [3,4,5,6,7] + [12,13,14,15,16]

def getTodayDinnerList(WD=MyShiftTime().weekday):
    WD = 0 if WD in [5,6] else WD
    
    wb = load_workbook(Filename)

    sheetnames = wb.get_sheet_names()
    sheetname = sheetnames[WD]
    sheet = wb[sheetname]

    Dinners = {'星期':sheetname,'餐點':list()}
    for i in cells:
        cell = 'A{}'.format(i)
        Dinner = sheet[cell].value
        Dinners['餐點'].append({cell:Dinner})

    return Dinners


##print(getTodayDinnerList(5))
